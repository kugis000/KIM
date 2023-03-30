import PySimpleGUI as sg
import os.path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference, LineChart, Series
import subprocess
import platform

# pārbauda vai excel files "dati.xlsx" jau eksiste, ja neeksistē tad izveido
filename = 'dati.xlsx'
if os.path.isfile(filename):
    wb = load_workbook(filename)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['Datums', 'Svars'])

#izdēš veco grafiku un izveido jauno
def update_chart_table(ws):
    # izdēš veco grafiku
    for chart in ws._charts:
        ws._charts.remove(chart)
    # izveido jauno grafiku
    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    dates = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data)
    chart.set_categories(dates)
    chart.title = "Svars laika gaitā"
    chart.x_axis.title = "Datums"
    chart.y_axis.title = "Svars(kg)"
    chart.legend = None #noņem leģendu
    simbols = chart.series[0]
    simbols.marker.symbol = "diamond"
    ws.add_chart(chart, 'G3')#upadato grafiku

kreisa = [
    [sg.Text("UZTURS", text_color="black", font=('Bahnschrift Condensed', 40), size =(30,1), justification='center')],
    [sg.Button('Olbaltumvielas', key = '-OLBALTUMVIELAS-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 20)),sg.Button('Ogļhidrāti', key = '-OGLHIDRATI-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 20)), sg.Button('Tauki', key = '-TAUKI-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 20)), sg.Button('Vitamīni', key = '-VITAMINI-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 20)), sg.Button('Minerālvielas', key = '-MINERALVIELAS-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 20))],
    [sg.Text('Ievadi svaru: '), sg.In(size=(25,1), enable_events=True, key='-INPUT1-')],
    [sg.Text('Ievadi garumu: '), sg.In(size=(25,1), enable_events=True, key='-INPUT2-')],
    [sg.Text('Ievadi vecumu: '), sg.In(size=(25,1), enable_events=True, key='-INPUT3-')],
    [sg.Text('Izvēlies aktivitātes līmeni: '), sg.InputCombo(('1.2', '1.375', '1.55', '1.725', '1.9'), size = (30,1), key = '-AKTIVITATE-', enable_events=True)],
    [sg.Button('ĶMI kalkulators', key = '-INDEKSS-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 15)), sg.Button('Pamata vielmaiņas ātruma kalkulators(BMR)', key = '-BMR-', size=(0, 0), visible=True, font=('Bahnschrift Condensed', 15))],
    [sg.Text('Uztura vēsture', text_color="dark blue", font=('Bahnschrift Condensed', 30), justification='center')],
    [sg.Text('•Vēsturiski uzturu ietekmēja cilvēka stāvoklis sabiedrībā, mazturīgo uzturs bieži bija nepilnīgs un nepietiekams, bet turīgo uzturs neveselīgs un pārmērīgs. Šādas diētas izraisīja dažādas slimības.', size=(80,2), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•Sākoties pasaules apceļošanai un jaunu vietu atklāšanai, pieauga uztura daudzveidība.', size=(80,1), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•Svarīgi pētījumi par uzturu tiek saistīti ar “Ķīmisko revolūciju” Francijā 18. gs. beigās, kad tika identificēti ķīmisko vielu pamatelementi un attīstījās ķīmisko analīžu metodes. Šie pētījumi turpinājās līdz 19. gs. 80.gadiem.', size=(80,3), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•Nākamie 25 gadi būtiski paplašināja izpratni par uzturvielu nepieciešamību, sasaistot nepietiekamu uzturvielu daudzumu ar dažādām slimībām.', size=(80,2), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•20. gs. sākums tiek dēvēts par vitamīnu pētniecības ēru, visā pasaulē tika veikti neskaitāmi pētījumi par vitamīniem un to ietekmi organismā.', size=(80,2), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•20. gs. vidū vitamīnu pētniecības ēru nomainīja olbaltumvielu pētniecība.', size=(80,1), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•20. gs. 60. gados konstatēja, ka olbaltumvielu nepietiekamība uzturā ir visnopietnākā un izplatītākā problēma pasaulē.', size=(80,2), font=('Bahnschrift Condensed', 15), justification='left')],
    [sg.Text('•Tad sākās taukvielu samazināšanas un ogļhidrātu palielināšanas kampaņas ASV, kas vēlāk pārņēma arī Eiropu, un cilvēkiem attīstījās vēl lielākas veselības problēmas, pārmērīga cukura patēriņa dēļ, attīstot diabēta un aptaukošanās problēmas.', size=(78,3), font=('Bahnschrift Condensed', 15), justification='left')]
]

laba = [
    [sg.Text('Šeit būs rezultāts',size=(80,1), font=('Bahnschrift Condensed', 15), justification='left', key='-rez-')]
]

layout=[
    [sg.Column(kreisa), sg.VSeparator(), sg.Column(laba)]
]
logs = sg.Window('Uzturs',layout,)
while True:
    event, values = logs.read()

    if event == sg.WIN_CLOSED:
        break

    if event == '-OLBALTUMVIELAS-':
        rezultats = "Olbaltumvielas ir svarīgas!"
        logs['-rez-']. update(rezultats)

    if event == '-OGLHIDRATI-':
        rezultats = "Oglhidrati ir svarīgi!"
        logs['-rez-']. update(rezultats)

    if event == '-TAUKI-':
        rezultats = "Tauki ir svarīgi!"
        logs['-rez-']. update(rezultats)

    if event == '-VITAMINI-':
        rezultats = "Vitamini ir vajadzigi!"
        logs['-rez-']. update(rezultats)

    if event == '-MINERALVIELAS-':
        rezultats = "Mineralvielas ir svarīgas!"
        logs['-rez-']. update(rezultats)

    if event == '-AKTIVITATE-':
        aktiv = float(values['-AKTIVITATE-'])

    if event == '-INDEKSS-':
        vert1 = values['-INPUT1-']
        vert2 = values['-INPUT2-']
        vert3 = values['-INPUT3-']
        aktiv = float(values['-AKTIVITATE-'])  # pārveifoju -AKTIVITATE- uz float
        saskaita = int(vert1) + int(vert2) + int(vert3) + aktiv
        rezultats = f"Saskaitot {vert1} un {vert2} un {vert3} un {aktiv} iegūst {saskaita}"
        logs['-rez-'].update(rezultats)

        # raksta tagadējo laiku un pievieno to worksheetam
        now = datetime.now().strftime('%Y-%m-%d')
        ws.append([now, float(vert1)])

        #updato grafiku
        update_chart_table(ws)

        #saglabā datus
        wb.save('dati.xlsx')

        # automātiski noskaidro operētājsistēmu un palaiž programmu,
        # kad jaunā inofrmācija ir saglabāta
        if platform.system() == 'Windows': # windows
            os.startfile('dati.xlsx')
        elif platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', 'dati.xlsx'))
        else:  # Linux, Unix, utt.
            subprocess.call(('xdg-open', 'dati.xlsx'))

    if event == '-BMR-':
        rezultats = "Mineālvielas ir vajadzīgas!"
        logs['-rez-']. update(rezultats)


logs.close()
